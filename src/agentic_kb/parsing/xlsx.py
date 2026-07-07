"""XLSX parser that preserves worksheet table and cell coordinates."""

import re
from collections.abc import Callable
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Any

from agentic_kb.parsing.base import (
    Parser,
    ParserDependencyError,
    ParserLimitError,
    ParserReadError,
    ParsingLimits,
    UnsupportedContentTypeError,
    raise_parser_read_error,
)
from agentic_kb.parsing.schemas import ParsedDocument, ParsedElement, ParsedSection


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
WorkbookLoader = Callable[[BytesIO], Any]
_AUTO_WORKBOOK_LOADER = object()


class XlsxParser(Parser):
    """Parse workbook sheets into table elements and row elements."""

    supported_content_types = (XLSX_CONTENT_TYPE,)

    def __init__(
        self,
        workbook_loader: WorkbookLoader | None | object = _AUTO_WORKBOOK_LOADER,
        limits: ParsingLimits | None = None,
    ) -> None:
        super().__init__(limits=limits)
        self._workbook_loader = workbook_loader

    def parse_text(
        self,
        text: str,
        *,
        source_uri: str,
        content_type: str,
    ) -> ParsedDocument:
        raise TypeError("XlsxParser requires bytes content")

    def parse(
        self,
        content: bytes | str,
        *,
        source_uri: str,
        content_type: str,
    ) -> ParsedDocument:
        if isinstance(content, str):
            content = content.encode("utf-8")
        self._ensure_supported(content_type)
        self._validate_content_size(content)
        normalized_content_type = content_type.split(";", 1)[0].strip().lower()

        workbook_loader = self._workbook_loader
        if workbook_loader is _AUTO_WORKBOOK_LOADER:
            workbook_loader = _default_workbook_loader()
        if workbook_loader is None:
            raise ParserDependencyError(
                "XLSX parsing requires openpyxl. Install agentic-kb[parsing] to enable XlsxParser."
            )

        try:
            workbook = workbook_loader(BytesIO(content))
            sections, elements = _extract_workbook(workbook)
        except (ParserDependencyError, ParserLimitError, ParserReadError, UnsupportedContentTypeError):
            raise
        except Exception as error:
            raise_parser_read_error(source_uri, normalized_content_type, error)
        finally:
            close = locals().get("workbook")
            if close is not None and hasattr(close, "close"):
                close.close()

        document = ParsedDocument(
            source_uri=source_uri,
            content_type=normalized_content_type,
            sections=sections,
            elements=elements,
            metadata={"sheet_count": len(sections), "table_count": len(sections)},
        )
        self._validate_document_limits(document)
        return document

    def _ensure_supported(self, content_type: str) -> None:
        normalized = content_type.split(";", 1)[0].strip().lower()
        if normalized not in self.supported_content_types:
            raise UnsupportedContentTypeError(f"XlsxParser does not support {content_type}")


def _default_workbook_loader() -> WorkbookLoader | None:
    """Resolve openpyxl lazily so workbook parsing remains an optional extra."""

    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    def load(stream: BytesIO) -> Any:
        return load_workbook(stream, read_only=True, data_only=True)

    return load


def _extract_workbook(workbook: Any) -> tuple[list[ParsedSection], list[ParsedElement]]:
    sections: list[ParsedSection] = []
    elements: list[ParsedElement] = []
    for sheet_index, worksheet in enumerate(workbook.worksheets):
        if getattr(worksheet, "sheet_state", "visible") != "visible":
            continue

        table = _table_from_sheet(worksheet)
        if table is None:
            continue

        table_text, columns, row_count, cell_range, row_elements = table
        # Apply post-parse pruning
        row_elements = _prune_row_elements(row_elements)
        row_count = len(row_elements)
        table_text = "\n".join(item["text"] for item in row_elements)
        section_path = (worksheet.title,)
        sections.append(
            ParsedSection(
                index=len(sections),
                title=worksheet.title,
                path=section_path,
                text=table_text,
                metadata={
                    "sheet_name": worksheet.title,
                    "sheet_index": sheet_index,
                    "row_count": row_count,
                    "column_count": len(columns),
                    "cell_range": cell_range,
                },
            )
        )
        elements.append(
            ParsedElement(
                index=len(elements),
                kind="table",
                text=table_text,
                section_path=section_path,
                metadata={
                    "sheet_name": worksheet.title,
                    "sheet_index": sheet_index,
                    "row_count": row_count,
                    "column_count": len(columns),
                    "columns": columns,
                    "cell_range": cell_range,
                },
            )
        )

    if sections:
        return sections, elements

    return [ParsedSection(index=0, title="XLSX Workbook", text="")], []


def _table_from_sheet(worksheet: Any) -> tuple[str, list[str], int, str, list[dict[str, Any]]] | None:
    rows = [
        tuple(row)
        for row in worksheet.iter_rows()
        if any(_cell_text(getattr(cell, "value", None)) for cell in row)
    ]
    if not rows:
        return None

    min_column = min(_first_non_empty_column(row) for row in rows)
    max_column = max(_last_non_empty_column(row) for row in rows)
    normalized_rows = [row[min_column - 1 : max_column] for row in rows]
    columns = _columns_from_header(normalized_rows[0], min_column)
    text_rows = [_row_text(row) for row in normalized_rows]
    row_elements = [
        _row_element(row, columns, min_column, max_column)
        for row in normalized_rows[1:]
    ]
    first_row = getattr(rows[0][min_column - 1], "row", 1)
    last_row = getattr(rows[-1][max_column - 1], "row", len(rows))
    cell_range = (
        f"{_column_letter(min_column)}{first_row}:"
        f"{_column_letter(max_column)}{last_row}"
    )
    return "\n".join(text_rows), columns, len(row_elements), cell_range, row_elements


def _row_element(
    row: tuple[Any, ...],
    columns: list[str],
    min_column: int,
    max_column: int,
) -> dict[str, Any]:
    values = [_cell_text(getattr(cell, "value", None)) for cell in row]
    # Normalize multi-line cell content to keep each row on a single line.
    values = [v.replace("\n", "; ").replace("\r", "") for v in values]
    row_number = getattr(row[0], "row", 0)
    # Prefix each cell value with its column header for retrieval clarity.
    prefixed = [
        f"{col}: {val}" if col and val else val
        for col, val in zip(columns, values)
    ]
    return {
        "text": " | ".join(prefixed),
        "metadata": {
            "row_number": row_number,
            "cell_range": (
                f"{_column_letter(min_column)}{row_number}:"
                f"{_column_letter(max_column)}{row_number}"
            ),
            "values": dict(zip(columns, values)),
            "cells": [
                {
                    "coordinate": getattr(cell, "coordinate", ""),
                    "column": column,
                    "value": value,
                }
                for cell, column, value in zip(row, columns, values)
                if not _is_empty_cell(cell)
            ],
        },
    }


def _columns_from_header(header: tuple[Any, ...], start_column: int) -> list[str]:
    columns: list[str] = []
    for offset, cell in enumerate(header):
        value = _cell_text(cell.value)
        if value:
            columns.append(value)
        else:
            columns.append(_column_letter(start_column + offset))
    return columns


def _row_text(row: tuple[Any, ...]) -> str:
    return " | ".join(_cell_text(cell.value) for cell in row)


def _first_non_empty_column(row: tuple[Any, ...]) -> int:
    for cell in row:
        if _is_empty_cell(cell):
            continue
        if _cell_text(getattr(cell, "value", None)):
            return getattr(cell, "column", 1)
    return getattr(row[0], "column", 1)


def _last_non_empty_column(row: tuple[Any, ...]) -> int:
    for cell in reversed(row):
        if _is_empty_cell(cell):
            continue
        if _cell_text(getattr(cell, "value", None)):
            return getattr(cell, "column", 1)
    return getattr(row[-1], "column", 1)


def _is_empty_cell(cell: Any) -> bool:
    """Check if a cell is openpyxl's EmptyCell type (no coordinate/row/column attrs)."""
    return type(cell).__name__ == "EmptyCell"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def _column_letter(column: int) -> str:
    letters = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


# ---- XLSX Row Pruning ----

_PAGE_NUMBER_PATTERN = re.compile(r"^(第\s*\d+\s*页|Page\s+\d+|-\s*\d+\s*-)$")
_SEPARATOR_PATTERN = re.compile(r"^[-=_]{3,}$")
_NOTE_MARKER = ("注：", "注意：", "说明：", "备注：", "Note:", "Notes:", "提示：")


def _prune_row_elements(row_elements: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Apply pruning filters to reduce noise in parsed XLSX rows."""

    if len(row_elements) <= 2:
        return row_elements

    pruned: list[dict[str, Any]] = []
    seen_text_signatures: set[str] = set()

    for element in row_elements:
        text = element["text"]
        values = element["metadata"].get("values", {})

        # 1. Drop totally empty rows
        if not any(str(v).strip() for v in values.values() if v):
            continue

        # 2. Drop separator / divider rows
        cells = [str(v).strip() for v in values.values() if v]
        if len(cells) == 1 and _SEPARATOR_PATTERN.match(cells[0]):
            continue

        # 3. Drop page-number / pagination rows
        if _is_page_number_row(cells):
            continue

        # 4. Drop note/annotation rows at the tail
        if _is_note_row(cells):
            continue

        # 5. Drop duplicate rows (same text content)
        text_sig = text.strip()
        if text_sig and text_sig in seen_text_signatures:
            continue
        seen_text_signatures.add(text_sig)

        # 6. Drop rows where every cell has the same value (e.g., "N/A", "-", "")
        unique_cells = {c for c in cells if c}
        if len(unique_cells) == 1 and len(unique_cells) < len(values):
            continue

        pruned.append(element)

    return pruned


def _is_page_number_row(cells: list[str]) -> bool:
    """Detect rows that are just page numbers or pagination markers."""
    non_empty = [c for c in cells if c]
    if len(non_empty) == 1:
        return bool(_PAGE_NUMBER_PATTERN.match(non_empty[0]))
    return False


def _is_note_row(cells: list[str]) -> bool:
    """Detect rows that are annotation/instruction markers."""
    if not cells:
        return False
    first = cells[0].strip()
    return any(first.startswith(marker) for marker in _NOTE_MARKER)
